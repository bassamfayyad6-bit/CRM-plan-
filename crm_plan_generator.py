import streamlit as st
import pandas as pd
import tempfile, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── colours ───────────────────────────────────────────────────────────────────
HEADER_BG  = '1F3864'; HEADER_FG  = 'FFFFFF'
LEGEND_BG  = '2E75B6'
SEC_FINAL  = 'E2EFDA'
SEC_PUSH   = 'FFF2CC'
SEC_INTANN = 'DDEEFF'
SEC_INTTRIM= 'EDE7F6'
SEC_NEW    = 'F5F5F5'
SEC_WARM   = 'FFE0B2'
NOTE_WR    = 'FFD700'
NOTE_SEC   = 'D6E4F0'
NOTE_END   = 'F4B183'

COLS = [
    ('NO',8),('COIL Man #',18),('A',7),('T.T',7),('TH [mm]',9),
    ('Width',8),('T.W',8),('Int+Trim',9),('Targeted Th.',12),
    ('Steel spool',11),('PASS',7),('Previous',14),('Process',10),
    ('NEXT',10),('Passes Left',11),
    ('Delivery date',14),('Notes',30),('Customer',28),
]
NCOLS = len(COLS)

SECTION_ORDER = ['FINAL', 'INT_ANN', 'PUSH', 'INT_TRIM']  # NEW goes to Sheet 2 only

SECTION_META = {
    'FINAL':    ('FINAL PASS COILS — 1 or 2 passes left, heading to F.Ann / T.L.L',
                 SEC_FINAL,  '1B5E20'),
    'INT_ANN':  ('INT ANNEALING COILS — next step: Intermediate Annealing',
                 SEC_INTANN, '1A3A5C'),
    'PUSH':     ('PUSH COILS — 3 passes left, clear path — roll 2 today / 1 tomorrow',
                 SEC_PUSH,   '7B5200'),
    'INT_TRIM': ('INT TRIM COILS — next step: Intermediate Trimming',
                 SEC_INTTRIM,'4A235A'),
    'NEW':      ('NEW COILS — P1 / P2 first passes',
                 SEC_NEW,    '3E3E3E'),
}

SEC_ROW_BG = {
    'FINAL': SEC_FINAL, 'INT_ANN': SEC_INTANN, 'PUSH': SEC_PUSH,
    'INT_TRIM': SEC_INTTRIM, 'NEW': SEC_NEW,
}

# ── helpers ───────────────────────────────────────────────────────────────────
def thin():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

def med_border(color='1F3864'):
    s = Side(style='medium', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def cs(ws, row, col, value, bold=False, bg=None, fg='000000',
       align='center', size=10, wrap=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Calibri', bold=bold, color=fg, size=size)
    if bg:
        c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = thin()
    if num_fmt:
        c.number_format = num_fmt
    return c

def banner(ws, r, text, bg, fg='000000', size=11, bdr='1F3864'):
    ws.row_dimensions[r].height = 24
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    c = ws.cell(row=r, column=1, value=text)
    c.font  = Font(name='Calibri', bold=True, color=fg, size=size)
    c.fill  = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = med_border(bdr)

def fmt_date(val):
    try:    return pd.Timestamp(val).strftime('%Y-%m-%d')
    except: return str(val) if pd.notna(val) else ''

def safe_str(val):
    return str(val).strip() if pd.notna(val) else ''

def safe_float(val):
    try:    return round(float(val), 3)
    except: return val

def del_ts(v):
    try:    return pd.Timestamp(v).timestamp()
    except: return 9e18

# ── data loading ──────────────────────────────────────────────────────────────
def load_data(filepath):
    master = pd.read_excel(filepath, sheet_name='Rolling Production Plan,', header=3)
    master.columns = [str(c).strip() for c in master.columns]
    crm = pd.read_excel(filepath, sheet_name='CRM ', header=2)
    crm.columns = [str(c).strip() for c in crm.columns]
    crm = crm[crm['COIL Man #'].notna() & crm['PASS'].notna()].copy()
    crm = crm[crm['PASS'].astype(str).str.startswith('P')].copy()
    # Dedup by coil+pass (keep first)
    crm['_key'] = crm['COIL Man #'].astype(str).str.strip() + '|' + crm['PASS'].astype(str).str.strip()
    crm = crm.drop_duplicates(subset=['_key'], keep='first').drop(columns=['_key'])
    # Separate P1/P2 from main queue
    is_p1p2 = crm['PASS'].astype(str).str.strip().isin(['P1', 'P2'])
    crm_main = crm[~is_p1p2].reset_index(drop=True)
    crm_new  = crm[is_p1p2].reset_index(drop=True)
    return master, crm_main, crm_new

# ── L2 integration ────────────────────────────────────────────────────────────

def load_l2(filepath):
    """
    Load L2 export file (XLS/XLSX).
    Returns DataFrame with last pass per coil (L2 is ground truth).
    Key columns: Coil No., Pass No., Exit Thickness [µm], End Time
    """
    try:
        df = pd.read_excel(filepath, engine='xlrd', header=0)
    except Exception:
        df = pd.read_excel(filepath, header=0)
    df.columns = [str(c).strip() for c in df.columns]

    # Normalize coil ID
    df['_coil_key'] = df['Coil No.'].astype(str).str.strip().str.upper()

    # Sort by End Time descending, keep last pass per coil
    df = df.sort_values('End Time', ascending=False)
    last = df.drop_duplicates(subset=['_coil_key'], keep='first').copy()
    last = last.set_index('_coil_key')
    return last

def apply_l2_to_crm(master, crm, l2_df):
    """
    Update CRM queue using L2 ground truth:
    1. For each coil in CRM, check if it exists in L2
    2. If L2 pass >= CRM pass → update to L2 actual state
    3. If coil finished (no more CRM passes after L2 pass) → remove from queue
    Returns updated crm DataFrame.
    """
    if l2_df is None or l2_df.empty:
        return crm

    updated_rows = []
    removed = []

    for _, r in crm.iterrows():
        coil_id  = safe_str(r['COIL Man #'])
        coil_key = coil_id.upper()

        if coil_key not in l2_df.index:
            # Not in L2 → keep as-is from plan
            updated_rows.append(r.to_dict())
            continue

        l2_row   = l2_df.loc[coil_key]
        l2_pass  = int(l2_row['Pass No.'])
        l2_th_um = l2_row['Exit Thickness [µm]']   # in micrometers
        l2_th_mm = round(float(l2_th_um) / 1000, 3) if pd.notna(l2_th_um) else None

        # Find what pass this coil should be on next in master
        j = get_master_journey(master, coil_id)
        if j.empty:
            updated_rows.append(r.to_dict())
            continue

        cm = j[j['PASS'].notna() & j['PASS'].astype(str).str.startswith('P')]

        # Find the NEXT pass after l2_pass in master
        next_passes = cm[cm['PASS'].astype(str).str.strip().apply(
            lambda p: int(p[1:]) if p[1:].isdigit() else 0) > l2_pass]

        if next_passes.empty:
            # No more CRM passes → coil is done or out of CRM
            removed.append(coil_id)
            continue

        # Next pass exists → update row with L2 actual state
        nxt = next_passes.iloc[0]
        updated = r.to_dict()
        updated['PASS']         = safe_str(nxt['PASS'])
        updated['TH [mm]']      = l2_th_mm if l2_th_mm else r['TH [mm]']
        updated['Targeted Th.'] = nxt.get('Targeted Th.', r.get('Targeted Th.'))
        updated['Process']      = safe_str(nxt.get('Process', r.get('Process', '')))
        updated['NEXT']         = safe_str(nxt.get('NEXT', r.get('NEXT', '')))
        updated['_l2_updated']  = True
        updated_rows.append(updated)

    if removed:
        print(f"  L2: removed {len(removed)} completed coils: {removed[:5]}{'...' if len(removed)>5 else ''}")

    result = pd.DataFrame(updated_rows).reset_index(drop=True).copy()
    # Ensure classification columns exist (will be computed in build_plan)
    # Remove any stale classification cols so build_plan recomputes them
    for col in ['_passes_left','_final_dest','_section','_del_sort']:
        if col in result.columns:
            result = result.drop(columns=[col])
    return result


# ── master lookup helpers ─────────────────────────────────────────────────────
def get_master_journey(master, coil_id):
    """Return all rows for this coil from master, sorted by NO."""
    j = master[master['COIL Man #'].astype(str).str.strip() ==
               str(coil_id).strip()].copy()
    return j.sort_values('NO') if not j.empty else pd.DataFrame()

def get_remaining_cm_passes(master, coil_id, cur_pass):
    """
    From master: count CRM passes (P-rows) remaining from cur_pass onward.
    Returns (count, final_NEXT) or (999, 'UNKNOWN') if not found.
    """
    j = get_master_journey(master, coil_id)
    if j.empty:
        return 999, 'UNKNOWN'
    cm = j[j['PASS'].notna() & j['PASS'].astype(str).str.startswith('P')]
    cur = cm[cm['PASS'].astype(str).str.strip() == str(cur_pass).strip()]
    if cur.empty:
        return 999, 'UNKNOWN'
    rem = cm[cm['NO'] >= cur.iloc[0]['NO']]
    return len(rem), safe_str(rem.iloc[-1]['NEXT'])

def get_previous_step(master, coil_id, cur_pass):
    """
    What delivered the coil to CRM for cur_pass?
    = NEXT value of the row immediately before cur_pass in the master journey.
    Fallback: P1 → HM, else → CM.
    """
    def fallback():
        return 'HM' if str(cur_pass).strip() == 'P1' else 'CM'

    j = get_master_journey(master, coil_id)
    if j.empty:
        return fallback()
    cm = j[j['PASS'].notna() & j['PASS'].astype(str).str.startswith('P')]
    cur = cm[cm['PASS'].astype(str).str.strip() == str(cur_pass).strip()]
    if cur.empty:
        return fallback()
    before = j[j['NO'] < cur.iloc[0]['NO']]
    if before.empty:
        return fallback()
    last = before.iloc[-1]
    nxt = safe_str(last.get('NEXT', ''))
    if nxt and nxt not in ('nan', 'None'):
        return nxt
    proc = safe_str(last.get('Process', ''))
    return proc if proc and proc not in ('nan', 'None') else fallback()

def has_int_step_remaining(master, coil_id, cur_pass):
    """
    True if the remaining journey from cur_pass onward contains ANY
    non-CM intermediate step (INT Ann, INT Trim, F Ann before final, etc.)
    i.e. the path is NOT: CM → CM → ... → F.Ann/T.L.L directly.
    We check the NEXT values of all remaining P-rows: if any is not CM
    and not the final F.Ann/T.L.L, path is not clear.
    """
    j = get_master_journey(master, coil_id)
    if j.empty:
        return True  # unknown → treat as not clear
    cm = j[j['PASS'].notna() & j['PASS'].astype(str).str.startswith('P')]
    cur = cm[cm['PASS'].astype(str).str.strip() == str(cur_pass).strip()]
    if cur.empty:
        return True
    rem = cm[cm['NO'] >= cur.iloc[0]['NO']]
    # All intermediate NEXTs (not the last one which is the final destination)
    intermediate_nexts = [safe_str(x).upper() for x in rem['NEXT'].tolist()[:-1]]
    # If any intermediate NEXT is not CM → not a clear path
    return any(n != 'CM' for n in intermediate_nexts)

# ── classify each CRM row ─────────────────────────────────────────────────────
def classify_row(master, r):
    """
    Classify a CRM row into a section based on:
    - NEXT column from CRM sheet (primary source)
    - Master plan for remaining passes count and path check
    """
    coil_id  = safe_str(r['COIL Man #'])
    cur_pass = safe_str(r['PASS'])
    nxt_crm  = safe_str(r.get('NEXT', '')).upper()  # NEXT from CRM sheet

    pl, final_dest = get_remaining_cm_passes(master, coil_id, cur_pass)

    # INT Ann: CRM says next step is INT Ann
    if 'INT' in nxt_crm and 'ANN' in nxt_crm:
        return 'INT_ANN', pl, final_dest

    # INT Trim: CRM says next step is INT Trim
    if 'INT' in nxt_crm and 'TRIM' in nxt_crm:
        return 'INT_TRIM', pl, final_dest

    # For CM→... paths, use master to determine how many passes and if path is clear
    # passes_left = passes AFTER current (not including current)
    if pl == 1:
        return 'FINAL', pl, final_dest

    if pl == 2:
        if not has_int_step_remaining(master, coil_id, cur_pass):
            return 'FINAL', pl, final_dest
        return 'NEW', pl, final_dest

    if pl == 3:
        if not has_int_step_remaining(master, coil_id, cur_pass):
            return 'PUSH', pl, final_dest
        return 'NEW', pl, final_dest

    return 'NEW', pl, final_dest

# ── warm-up selection ─────────────────────────────────────────────────────────
def select_warmup(master, crm, used_coils, n=3, min_width=0):
    """
    Warm-up coils from CRM sheet:
    - TH [mm] between 2.0 and 3.0
    - NEXT = INT Trim or INT Ann (from CRM sheet)
    - All selected must share the same NEXT destination
    Pick group with more candidates; tie → earliest delivery date.
    """
    int_ann, int_trim = [], []

    for _, r in crm.iterrows():
        coil_id = safe_str(r['COIL Man #'])
        if coil_id in used_coils:
            continue
        try:
            th = float(r.get('TH [mm]', 0))
        except:
            continue
        if not (2.0 <= th <= 3.0):
            continue
        try:
            w = float(r.get('Width', 0))
        except:
            w = 0
        if w <= min_width:
            continue
        nxt = safe_str(r.get('NEXT', '')).upper()
        pl, fd = get_remaining_cm_passes(master, coil_id, safe_str(r['PASS']))
        entry = {**r.to_dict(), '_passes_left': pl, '_final_dest': fd,
                 '_del_sort': del_ts(r.get('Delivery date'))}
        if 'INT' in nxt and 'ANN' in nxt:
            int_ann.append(entry)
        elif 'INT' in nxt and 'TRIM' in nxt:
            int_trim.append(entry)

    if not int_ann and not int_trim:
        return pd.DataFrame()
    group = int_ann if len(int_ann) >= len(int_trim) else int_trim
    return (pd.DataFrame(group).sort_values('Width', ascending=False)
              .head(n).reset_index(drop=True))

# ── build next-pass synthetic row from master ─────────────────────────────────
def make_next_pass_row(master, base_row):
    """
    Given a CRM row, look up the NEXT CRM pass in master and return a
    synthetic row representing that pass (for pairs in FINAL and PUSH sections).
    Returns None if not found.
    """
    coil_id  = safe_str(base_row['COIL Man #'])
    cur_pass = safe_str(base_row['PASS'])

    j = get_master_journey(master, coil_id)
    if j.empty:
        return None
    cm = j[j['PASS'].notna() & j['PASS'].astype(str).str.startswith('P')]
    cur = cm[cm['PASS'].astype(str).str.strip() == cur_pass]
    if cur.empty:
        return None
    nxt_rows = cm[cm['NO'] > cur.iloc[0]['NO']]
    if nxt_rows.empty:
        return None

    nxt = nxt_rows.iloc[0]
    pl_nxt, fd_nxt = get_remaining_cm_passes(master, coil_id, safe_str(nxt['PASS']))

    d = base_row.to_dict()
    d['PASS']          = safe_str(nxt['PASS'])
    d['TH [mm]']       = nxt.get('TH [mm]', '')
    d['Targeted Th.']  = nxt.get('Targeted Th.', '')
    d['Process']       = safe_str(nxt.get('Process', ''))
    d['NEXT']          = safe_str(nxt.get('NEXT', ''))
    d['_passes_left']  = pl_nxt
    d['_final_dest']   = fd_nxt
    d['_is_synthetic'] = True
    return pd.Series(d)

# ── build plan ────────────────────────────────────────────────────────────────
def build_plan(master, crm):  # crm here is crm_main (no P1/P2)
    # Step 1: classify every CRM row
    classified = []
    for _, r in crm.iterrows():
        sec, pl, fd = classify_row(master, r)
        classified.append({**r.to_dict(),
                            '_section':    sec,
                            '_passes_left': pl,
                            '_final_dest': fd,
                            '_del_sort':   del_ts(r.get('Delivery date'))})
    df = pd.DataFrame(classified)

    # Step 2: warm-up from CRM (TH 2-3, INT dest, same group)
    # Warm-up coils must be wider than the widest FINAL coil
    final_rows_all = [r for r in classified if r['_section'] == 'FINAL']
    max_final_width = max((float(r.get('Width', 0)) for r in final_rows_all
                           if pd.notna(r.get('Width'))), default=0)
    df_warmup = select_warmup(master, crm, set(), n=3, min_width=max_final_width)
    warmup_keys = set()
    if not df_warmup.empty:
        warmup_keys = set(
            (safe_str(r['COIL Man #']) + '|' + safe_str(r['PASS']))
            for _, r in df_warmup.iterrows()
        )

    # Remove warm-up coils from main pool
    df['_key'] = df['COIL Man #'].astype(str).str.strip() + '|' + df['PASS'].astype(str).str.strip()
    df = df[~df['_key'].isin(warmup_keys)].copy()

    # Step 3: build each section, tracking placed keys to avoid duplicates
    placed_keys = set()
    sections = {}

    # FINAL: pl==1 direct + pl==2 clear path as pairs
    final_rows = []
    df_final = df[df['_section'] == 'FINAL']
    finals_2 = df_final[df_final['_passes_left'] == 2].sort_values('Width', ascending=False)
    finals_1 = df_final[df_final['_passes_left'] == 1]
    paired_coils = set(finals_2['COIL Man #'].astype(str).str.strip())
    finals_1 = finals_1[~finals_1['COIL Man #'].astype(str).str.strip().isin(paired_coils)]

    for _, r in finals_2.iterrows():
        k = safe_str(r['COIL Man #']) + '|' + safe_str(r['PASS'])
        if k in placed_keys: continue
        final_rows.append(r)
        placed_keys.add(k)
        nxt_row = make_next_pass_row(master, r)
        if nxt_row is not None:
            nxt_k = safe_str(nxt_row['COIL Man #']) + '|' + safe_str(nxt_row['PASS'])
            if nxt_k not in placed_keys:
                final_rows.append(nxt_row)
                placed_keys.add(nxt_k)

    for _, r in finals_1.sort_values('Width', ascending=False).iterrows():
        k = safe_str(r['COIL Man #']) + '|' + safe_str(r['PASS'])
        if k in placed_keys: continue
        final_rows.append(r)
        placed_keys.add(k)

    sections['FINAL'] = pd.DataFrame(final_rows).reset_index(drop=True) if final_rows else pd.DataFrame()

    # PUSH: pl==3 clear path as pairs (cur + next pass)
    push_rows = []
    for _, r in df[df['_section'] == 'PUSH'].sort_values('Width', ascending=False).iterrows():
        k = safe_str(r['COIL Man #']) + '|' + safe_str(r['PASS'])
        if k in placed_keys: continue
        push_rows.append(r)
        placed_keys.add(k)
        nxt_row = make_next_pass_row(master, r)
        if nxt_row is not None:
            nxt_k = safe_str(nxt_row['COIL Man #']) + '|' + safe_str(nxt_row['PASS'])
            if nxt_k not in placed_keys:
                push_rows.append(nxt_row)
                placed_keys.add(nxt_k)

    sections['PUSH'] = pd.DataFrame(push_rows).reset_index(drop=True) if push_rows else pd.DataFrame()

    # INT_ANN, INT_TRIM, NEW
    for sec in ['INT_ANN', 'INT_TRIM', 'NEW']:
        rows = []
        for _, r in df[df['_section'] == sec].sort_values('Width', ascending=False).iterrows():
            k = safe_str(r['COIL Man #']) + '|' + safe_str(r['PASS'])
            if k in placed_keys: continue
            rows.append(r)
            placed_keys.add(k)
        sections[sec] = pd.DataFrame(rows).reset_index(drop=True) if rows else pd.DataFrame()

    return df_warmup, sections

# ── write one data row ────────────────────────────────────────────────────────
def write_row(ws, excel_row, no_val, row, bg, master):
    ws.row_dimensions[excel_row].height = 17

    def w(col, val, num_fmt=None):
        cs(ws, excel_row, col, val, bg=bg, num_fmt=num_fmt)

    w(1,  no_val)
    w(2,  safe_str(row.get('COIL Man #', '')))
    w(3,  row.get('A', ''))
    w(4,  row.get('T.T', ''))
    w(5,  safe_float(row.get('TH [mm]', '')), '0.000')
    w(6,  row.get('Width', ''))
    w(7,  row.get('T.W', ''))
    w(8,  row.get('Int + Final Trim', ''))
    w(9,  safe_float(row.get('Targeted Th.', '')), '0.000')
    w(10, row.get('Steel spool', ''))
    w(11, safe_str(row.get('PASS', '')))

    # Previous: from CRM sheet first, then master lookup
    prev_val = safe_str(row.get('Previous', ''))
    if not prev_val or prev_val in ('nan', 'None'):
        prev_val = get_previous_step(master, row.get('COIL Man #', ''), row.get('PASS', ''))
    w(12, prev_val)

    w(13, safe_str(row.get('Process', '')))
    w(14, safe_str(row.get('NEXT', '')) if pd.notna(row.get('NEXT')) else '')

    pl = row.get('_passes_left', 999) if isinstance(row, dict) else (row['_passes_left'] if '_passes_left' in row.index else 999)
    pl_c = ws.cell(row=excel_row, column=15, value=pl if pl < 999 else 'N/A')
    pl_c.font      = Font(name='Calibri', bold=(pl >= 999),
                          color=('C00000' if pl >= 999 else '000000'), size=10)
    pl_c.fill      = PatternFill('solid', start_color=bg)
    pl_c.alignment = Alignment(horizontal='center', vertical='center')
    pl_c.border    = thin()

    w(16, fmt_date(row.get('Delivery date')))

    notes_val = safe_str(row.get('Notes', ''))
    nc = ws.cell(row=excel_row, column=17, value=notes_val)
    nc.font      = Font(name='Calibri', size=9,
                        color='C00000' if 'ANN' in notes_val.upper() else '000000')
    nc.fill      = PatternFill('solid', start_color=bg)
    nc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    nc.border    = thin()

    cc = ws.cell(row=excel_row, column=18, value=safe_str(row.get('Customer', '')))
    cc.font      = Font(name='Calibri', size=9)
    cc.fill      = PatternFill('solid', start_color=bg)
    cc.alignment = Alignment(horizontal='left', vertical='center')
    cc.border    = thin()

# ── build Excel ───────────────────────────────────────────────────────────────
def build_excel(df_warmup, sections, plan_date, master, section_order=None, new_coils=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'CRM Plan'

    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    tc = ws.cell(row=1, column=1,
                 value=f'CRM Production Plan  —  {plan_date}  |  WR check @ pass 110  |  WR change @ pass 150')
    tc.font      = Font(name='Calibri', bold=True, color=HEADER_FG, size=14)
    tc.fill      = PatternFill('solid', start_color=HEADER_BG)
    tc.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[2].height = 15
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    lc = ws.cell(row=2, column=1,
                 value=('🟢 Final Pass (1-2 passes left → F.Ann / T.L.L)   '
                        '🟡 Push Coils (3 passes left, clear path — 2 today / 1 tomorrow)   '
                        '🔵 INT Ann   🟣 INT Trim   🟠 Warm-up  |  ⬜ P1/P2 → Sheet 2'))
    lc.font      = Font(name='Calibri', italic=True, color=HEADER_FG, size=9)
    lc.fill      = PatternFill('solid', start_color=LEGEND_BG)
    lc.alignment = Alignment(horizontal='center', vertical='center')

    HDR = 3
    ws.row_dimensions[HDR].height = 22
    for ci, (name, width) in enumerate(COLS, 1):
        cs(ws, HDR, ci, name, bold=True, bg=HEADER_BG, fg=HEADER_FG)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = f'A{HDR + 1}'

    cur_row  = HDR + 1
    pass_cnt = 0
    MAX      = 999999  # No limit — process all coils except P1/P2

    # warm-up
    banner(ws, cur_row,
           '⚠️   CHANGE WORK ROLL   —   Warm up the mill with the coils below   ⚠️',
           NOTE_WR, '7B0000', size=12, bdr='BF8600')
    cur_row += 1

    if not df_warmup.empty:
        for _, wrow in df_warmup.iterrows():
            write_row(ws, cur_row, f'W{pass_cnt+1}', wrow, SEC_WARM, master)
            cur_row += 1; pass_cnt += 1
    else:
        banner(ws, cur_row, 'No warm-up candidates found — select manually',
               SEC_WARM, 'C00000', size=10)
        cur_row += 1

    # sections
    order = section_order if section_order else SECTION_ORDER
    for sec in order:
        if pass_cnt >= MAX: break
        sub = sections.get(sec, pd.DataFrame())
        if sub.empty: continue
        label, _, _ = SECTION_META[sec]
        row_bg_col  = SEC_ROW_BG[sec]
        banner(ws, cur_row, f'▶   {label}', NOTE_SEC, '1F3864', size=10, bdr='2E75B6')
        cur_row += 1
        for _, row in sub.iterrows():
            write_row(ws, cur_row, pass_cnt + 1, row, row_bg_col, master)
            cur_row += 1; pass_cnt += 1
            # At pass 110: check work roll banner
            if pass_cnt == 110:
                banner(ws, cur_row,
                       '🔧   PASS 110  —  CHECK WORK ROLL CONDITION  —  UPDATE PLAN IF NEEDED   🔧',
                       'FFF3CD', '7B5200', size=11, bdr='BF8600')
                cur_row += 1
            # At pass 150: change work roll banner
            if pass_cnt == 150:
                banner(ws, cur_row,
                       '⚠️   PASS 150  —  CHANGE WORK ROLL  —  UPDATE THE PLAN   ⚠️',
                       NOTE_WR, '7B0000', size=12, bdr='BF8600')
                cur_row += 1

    banner(ws, cur_row,
           f'📋   PLAN COMPLETE  —  {pass_cnt} passes total  —  See Sheet 2 for New Coils (P1/P2)   📋',
           NOTE_END, '7B0000', size=12, bdr='BF6000')

    ws.auto_filter.ref = f'A{HDR}:{get_column_letter(NCOLS)}{cur_row}'

    # ── Sheet 2: New Coils (P1/P2) ───────────────────────────────────────
    if new_coils is not None and not new_coils.empty:
        ws2 = wb.create_sheet(title='New Coils (P1-P2)')
        ws2.row_dimensions[1].height = 28
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
        tc2 = ws2.cell(row=1, column=1,
                       value=f'New Coils — P1 / P2  |  {plan_date}  (use if main plan is exhausted)')
        tc2.font      = Font(name='Calibri', bold=True, color=HEADER_FG, size=13)
        tc2.fill      = PatternFill('solid', start_color='3E3E3E')
        tc2.alignment = Alignment(horizontal='center', vertical='center')

        ws2.row_dimensions[2].height = 22
        for ci, (name, width) in enumerate(COLS, 1):
            cs(ws2, 2, ci, name, bold=True, bg=HEADER_BG, fg=HEADER_FG)
            ws2.column_dimensions[get_column_letter(ci)].width = width
        ws2.freeze_panes = 'A3'

        for idx, (_, row) in enumerate(new_coils.iterrows()):
            er = 3 + idx
            ws2.row_dimensions[er].height = 17
            write_row(ws2, er, idx + 1, row, SEC_NEW, master)

        ws2.auto_filter.ref = f'A2:{get_column_letter(NCOLS)}{2 + len(new_coils)}'

    return wb

# ── Streamlit UI ──────────────────────────────────────────────────────────────
st.set_page_config(page_title='CRM Plan Generator', page_icon='🏭', layout='centered')
st.title('🏭 CRM Plan Generator')

uploaded = st.file_uploader('Upload Full Schedule (Cold Rolling Schedule)', type=['xlsx'])

col_l2, col_info = st.columns([2,1])
with col_l2:
    l2_file = st.file_uploader('Upload L2 Export (last 4 days) — optional but recommended', 
                                type=['xls','xlsx'],
                                help='L2 data overrides the plan for coils already processed')
with col_info:
    if l2_file:
        st.success('✅ L2 loaded — plan will be updated with actual mill data')
    else:
        st.info('ℹ️ Without L2, plan is based on schedule only')

# Section labels for display
SEC_LABELS = {
    'FINAL':    '🟢 Final Pass',
    'INT_ANN':  '🔵 INT Annealing',
    'PUSH':     '🟡 Push Coils',
    'INT_TRIM': '🟣 INT Trim',
}

st.markdown('### Priority Order')
st.caption('Drag to reorder — warm-up coils are always first (fixed)')

# Default order
DEFAULT_ORDER = ['FINAL', 'INT_ANN', 'PUSH', 'INT_TRIM']  # NEW always goes to Sheet 2

# Build priority selector using selectboxes
col1, col2 = st.columns([1, 2])
with col1:
    st.markdown('**Position**')
    for i in range(1, 5):
        st.markdown(f'**{i}.**')
        st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

with col2:
    st.markdown('**Section**')
    selected_order = []
    remaining_options = list(DEFAULT_ORDER)

    for i in range(4):
        default_idx = i if i < len(remaining_options) else 0
        choice = st.selectbox(
            f'Position {i+1}',
            options=remaining_options,
            index=default_idx,
            format_func=lambda x: SEC_LABELS[x],
            key=f'sec_{i}',
            label_visibility='collapsed'
        )
        selected_order.append(choice)
        remaining_options = [x for x in DEFAULT_ORDER if x not in selected_order]

# Validate no duplicates (fill any missing)
final_order = []
seen = set()
for s in selected_order:
    if s not in seen:
        final_order.append(s)
        seen.add(s)
for s in DEFAULT_ORDER:
    if s not in seen:
        final_order.append(s)
        seen.add(s)

# Show preview
st.markdown('**Plan order:** 🟠 Warm-up → ' + ' → '.join(SEC_LABELS[s] for s in final_order) + '  |  ⬜ P1/P2 → Sheet 2')

st.divider()

if uploaded:
    with st.spinner('Processing...'):
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                tmp.write(uploaded.read())
                tmp_path = tmp.name

            master, crm, crm_p1p2 = load_data(tmp_path)

            # Apply L2 if uploaded
            l2_df = None
            if l2_file:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as tmp_l2:
                    tmp_l2.write(l2_file.read())
                    tmp_l2_path = tmp_l2.name
                l2_df = load_l2(tmp_l2_path)
                crm = apply_l2_to_crm(master, crm, l2_df)
                os.unlink(tmp_l2_path)
                l2_updated = crm.get('_l2_updated', pd.Series(dtype=bool)).sum() if '_l2_updated' in crm.columns else 0
                st.caption(f'L2 applied: {len(l2_df)} coils from L2, {l2_updated} updated in plan')

            df_warmup, sections = build_plan(master, crm)

            # Apply user-selected order
            total     = sum(len(v) for v in sections.values())
            plan_date = datetime.today().strftime('%Y-%m-%d')
            # P1/P2 come from crm_p1p2 (separated at load time)
            new_coils = crm_p1p2.copy()
            new_coils['_passes_left'] = 999
            new_coils['_final_dest']  = ''
            new_coils['_section']     = 'NEW'
            new_coils['_del_sort']    = 0
            new_coils['_is_synthetic']= False
            new_coils = new_coils.sort_values('Width', ascending=False).reset_index(drop=True)
            wb        = build_excel(df_warmup, sections, plan_date, master,
                                    section_order=final_order,
                                    new_coils=new_coils)
            out_path  = os.path.join(tempfile.gettempdir(),
                                     f'CRM_Plan_{plan_date}.xlsx')
            wb.save(out_path)
            with open(out_path, 'rb') as f:
                excel_bytes = f.read()
            os.unlink(tmp_path)

            # Show counts
            cols = st.columns(4)
            for i, sec in enumerate(final_order):
                df_sec = sections.get(sec, pd.DataFrame())
                cols[i].metric(SEC_LABELS[sec], len(df_sec))

            st.download_button(
                label='⬇️  Download CRM Plan',
                data=excel_bytes,
                file_name=f'CRM_Plan_{plan_date}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
                type='primary'
            )
        except Exception as e:
            st.error(f'Error: {e}')
            import traceback
            st.code(traceback.format_exc())
